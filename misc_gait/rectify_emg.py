# -*- coding: utf-8 -*-
"""
Read EMG channels from Nexus. Calculate linear envelope and rectified signal.
Downsample and write as model outputs.

Based on extractEMG.py by Ned Pires.

@author: Jussi (jnu@iki.fi)

requires: gaitutils, numpy, scipy
"""

# %% init

import os.path as op
import numpy as np
import scipy
import logging
import openpyxl
from openpyxl.styles import Font
import matplotlib.pyplot as plt

from gaitutils import nexus, sessionutils, cfg, trial

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)


def _compute_emg_envelope():
    """Compute EMG linear envelope and rectified signal for currently open Nexus trial.

    Write as model outputs.
    """

    # define parameters
    HPF = 5  # high pass frequency
    LPF = 10  # low pass frequency
    BUTTER_ORDER = 4  # filter order

    # read EMG data
    vicon = nexus.viconnexus()
    emgdata = nexus._get_emg_data(vicon)['data']
    meta = nexus._get_metadata(vicon)
    emgrate = meta['analograte']
    subject = meta['name']
    nframes = meta['length']

    # apply hpf
    b_HPF, a_HPF = scipy.signal.butter(
        BUTTER_ORDER, HPF * 2 / emgrate, 'high', analog=False
    )
    emg_hpf = {
        chname: scipy.signal.filtfilt(b_HPF, a_HPF, chdata)
        for chname, chdata in emgdata.items()
    }

    # rectify
    emg_rectified = {chname: np.abs(chdata) for chname, chdata in emg_hpf.items()}

    # apply lpf
    b_LPF, a_LPF = scipy.signal.butter(
        BUTTER_ORDER, LPF * 2 / emgrate, 'low', analog=False
    )
    emg_rectified_lpf = {
        chname: scipy.signal.filtfilt(b_LPF, a_LPF, chdata)
        for chname, chdata in emg_rectified.items()
    }

    # downsample
    emg_rectified_ds = {
        chname + '_Rectified': scipy.signal.resample(chdata, nframes)
        for chname, chdata in emg_rectified.items()
    }
    emg_linearenvelope_ds = {
        chname + '_LinearEnvelope': scipy.signal.resample(chdata, nframes)
        for chname, chdata in emg_rectified_lpf.items()
    }

    # create the new model outputs in Nexus
    existing_outputs = vicon.GetModelOutputNames(subject)
    new_outputs = emg_rectified_ds.keys() + emg_linearenvelope_ds.keys()
    for output in set(new_outputs) - set(existing_outputs):
        logger.debug('creating model output %s' % output)
        vicon.CreateModelOutput(
            subject,
            output,
            'EMG',
            {'EMG', 'EMG', 'EMG'},
            {'Electric Potential', 'Electric Potential', 'Electric Potential'},
        )

    # write the processed EMG as model outputs
    exists = [True] * nframes

    for chname, chdata in emg_rectified_ds.items():
        logger.debug('writing data for %s' % chname)
        vicon.SetModelOutput(subject, chname, [chdata], exists)

    for chname, chdata in emg_linearenvelope_ds.items():
        logger.debug('writing data for %s' % chname)
        vicon.SetModelOutput(
            subject,
            chname,
            [chdata],
            exists,
        )

    return new_outputs


def _get_model_output(vicon, subj, var):
    """Read a single model output from Nexus"""
    nums, _ = vicon.GetModelOutput(subj, var)
    if nums:
        data = np.squeeze(np.array(nums))
    else:
        logger.info('cannot read model variable %s, returning nans' % var)
        data = None
    return data


def _bold_cell(ws, **cell_params):
    """Write a bold-styled cell into worksheet ws"""
    boldfont = Font(bold=True)
    _cell = ws.cell(**cell_params)
    _cell.font = boldfont


def _guess_emg_devname(vicon):
    """Try to guess the EMG device name"""
    devnames = ['Myon EMG', 'Noraxon Ultium']  # the candidates
    for devname in devnames:
        ids = [
            id_
            for id_ in vicon.GetDeviceIDs()
            if vicon.GetDeviceDetails(id_)[0].lower() == devname.lower()
        ]
        if ids:
            return devname


# %%
if __name__ == '__main__':
    # loop through session c3d files, compute envelopes and save into c3ds and xlsx
    cfg.autoproc.nexus_forceplate_devnames = []  # read all forceplates
    vicon = nexus.viconnexus()
    subj = nexus.get_subjectnames()
    emg_devname = _guess_emg_devname(vicon)
    if emg_devname is None:
        raise RuntimeError('Cannot figure out EMG device name')
    cfg.emg.devname = emg_devname
    # get and process all session trials
    sp = nexus.get_sessionpath()
    c3ds = sessionutils.get_c3ds(sp, tags=['XXX'])
    norm_data = dict()
    fname_xls = op.join(sp, op.split(sp)[-1] + '.xlsx')
    wb = openpyxl.Workbook()
    for n, c3dfile in enumerate(c3ds[:2]):
        ncycles = dict()
        logger.debug('opening %s' % c3dfile)
        nexus._open_trial(c3dfile)
        # compute the envelopes into Nexus model vars; names are returned
        modelvars = _compute_emg_envelope()
        tr = trial.nexus_trial()
        for ctxt in 'LR':
            this_cycles = tr.get_cycles({ctxt: 'all'})
            if ctxt not in ncycles:
                ncycles[ctxt] = len(this_cycles)
            this_vars = [var for var in modelvars if var[0] == ctxt]
            for var in this_vars:
                norm_data[var] = list()
                for cyc in this_cycles:
                    this_data = _get_model_output(vicon, subj, var)
                    this_data_norm = cyc.normalize(this_data)[1]
                    norm_data[var].append(this_data_norm)
                norm_data[var] = np.array(norm_data[var])
        avg_data = dict()
        std_data = dict()
        for var in modelvars:
            avg_data[var] = norm_data[var].mean(axis=0)
            std_data[var] = norm_data[var].std(axis=0)
        ws = wb.active if n == 0 else wb.create_sheet()
        ws.title = op.splitext(op.split(c3dfile)[-1])[0]
        # write some metadata
        _bold_cell(ws, column=1, row=2, value='Subject:')
        _bold_cell(ws, column=1, row=3, value='Trial:')
        _bold_cell(ws, column=1, row=4, value='Averaged cycles left:')
        _bold_cell(ws, column=1, row=5, value='Averaged cycles right:')
        ws.cell(column=2, row=2, value=subj)
        ws.cell(column=2, row=3, value=op.split(c3dfile)[-1])
        ws.cell(column=2, row=4, value=ncycles['L'])
        ws.cell(column=2, row=5, value=ncycles['R'])
        # write channel avg and std data
        for col, var in enumerate(sorted(modelvars), 3):
            _bold_cell(ws, column=2 * col, row=2, value=var + ' mean')  # col header
            _bold_cell(
                ws, column=2 * col + 1, row=2, value=var + ' stddev'
            )  # col header
            for row, x in enumerate(avg_data[var], 3):
                ws.cell(column=2 * col, row=row, value=x)
            for row, x in enumerate(std_data[var], 3):
                ws.cell(column=2 * col + 1, row=row, value=x)
    wb.save(filename=fname_xls)
