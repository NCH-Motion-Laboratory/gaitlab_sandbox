# -*- coding: utf-8 -*-
"""

Rectify EMG from c3d files. Average and write into XLSX.

@author: Jussi (jnu@iki.fi)

"""

# %% init

import os.path as op
import numpy as np
from numpy.core.fromnumeric import _choose_dispatcher
import scipy
import logging
import openpyxl
from openpyxl.styles import Font
import matplotlib.pyplot as plt
from collections import defaultdict

from gaitutils.envutils import GaitDataError
from gaitutils.numutils import _isint
from gaitutils import c3d, nexus, sessionutils, cfg, trial, read_data

logging.basicConfig(level=logging.WARNING)
logger = logging.getLogger(__name__)



def _bold_cell(ws, **cell_params):
    """Write a bold-styled cell into worksheet ws"""
    boldfont = Font(bold=True)
    _cell = ws.cell(**cell_params)
    _cell.font = boldfont




def _compute_emg_envelope_c3d(c3dfile):
    """Compute EMG linear envelope and rectified signal for a c3d file"""

    # define parameters
    HPF = 5  # high pass frequency
    LPF = 10  # low pass frequency
    BUTTER_ORDER = 4  # filter order

    # read EMG data
    emgdata = read_data.get_emg_data(c3dfile)['data']
    meta = read_data.get_metadata(c3dfile)
    emgrate = meta['analograte']
    nframes = meta['length']

    # strip Voltage. prefix that Nexus inserts
    emgdata = {(chname[8:] if chname.find('Voltage') == 0 else chname): data for chname, data in emgdata.items()}

    # apply hpf
    b_HPF, a_HPF = scipy.signal.butter(
        BUTTER_ORDER, HPF * 2 / emgrate, 'high', analog=False
    )
    emg_hpf = {
        chname: scipy.signal.filtfilt(b_HPF, a_HPF, chdata)
        for chname, chdata in emgdata.items()
    }

    # rectify
    emg_rectified = {chname: np.abs(chdata) for chname, chdata in emg_hpf.items()}

    # apply lpf
    b_LPF, a_LPF = scipy.signal.butter(
        BUTTER_ORDER, LPF * 2 / emgrate, 'low', analog=False
    )
    emg_rectified_lpf = {
        chname: scipy.signal.filtfilt(b_LPF, a_LPF, chdata)
        for chname, chdata in emg_rectified.items()
    }

    # downsample
    emg_rectified_ds = {
        chname + '_Rectified': scipy.signal.resample(chdata, nframes)
        for chname, chdata in emg_rectified.items()
    }
    emg_linearenvelope_ds = {
        chname + '_LinearEnvelope': scipy.signal.resample(chdata, nframes)
        for chname, chdata in emg_rectified_lpf.items()
    }

    return emg_linearenvelope_ds


def _channel_context(chname, idx_mapper):
    """Try to figure out channel context.
    
    idx_mapper must be a dict that maps channel index to context.
    """
    if chname[0] in 'LR':  # Myon naming
        return chname[0]
    elif chname[:3] == 'EMG':  # Noraxon naming
        dot_pos = chname.find('.')
        idx = chname[3:dot_pos]  # channel index
        try:
            return idx_mapper(int(idx))
        except ValueError:
            raise ValueError('cannot parse channel name %s' % chname)
    else:  # unrecognized channel
        return None


def _auto_adjust(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value


# %% read through EMG, compute envelopes, save into XLSX


# emg1-6 oikea, paitsi Vilma emg1-6 vasen

# Noraxon-specific mapping from ch index to context
idx_mapper = lambda idx: 'R' if idx <= 6 else 'L'
idx_mapper_reverse = lambda idx: 'L' if idx <= 6 else 'R'

cfg.autoproc.nexus_forceplate_devnames = []  # read all forceplates

session_root = r'C:\Users\hus20664877\Downloads\C3D files'

fname_xls = op.join(session_root, 'emg_envelopes.xlsx')

norm_data = dict()
wb = openpyxl.Workbook()

# be more tolerant about toeoffs
cfg.trial.no_toeoff = 'reject'
cfg.trial.multiple_toeoffs = 'reject'

# get the c3ds
allfiles = list()
for d0, dirs, files in os.walk(session_root):
    allfiles.extend(op.join(d0, fn) for fn in files if '.c3d' in fn.lower())

# compute envelopes, normalize to cycles (for matching context) and average (one per trial)
for n, c3dfile in enumerate(sorted(allfiles)):
    try:
        lenv = _compute_emg_envelope_c3d(c3dfile)
    except GaitDataError:
        logger.warning('cannot read EMG from %s, skipping' % c3dfile)
        continue
    tr = trial.Trial(c3dfile)
    this_cycles = tr.get_cycles('all')
    # count L/R cycles
    ncycles = {ctxt: len([c for c in this_cycles if c.context == ctxt]) for ctxt in 'LR'}
    # for Vilma, we have a different channel mapping
    _idx_mapper = idx_mapper_reverse if 'VILMA' in c3dfile.upper() else idx_mapper
    norm_data = dict()
    chs_valid = [ch for ch in lenv if _channel_context(ch, _idx_mapper) is not None]
    for chname in chs_valid:
        chdata = lenv[chname]
        ctxt_this = _channel_context(chname, idx_mapper)
        norm_data[chname] = list()
        for cyc in this_cycles:
            if ctxt_this != cyc.context:
                continue
            chdata_norm = cyc.normalize(chdata)[1]
            norm_data[chname].append(chdata_norm)
        norm_data[chname] = np.array(norm_data[chname])

    avg_data = dict()
    std_data = dict()
    for chname in chs_valid:
        avg_data[chname] = norm_data[chname].mean(axis=0)
        std_data[chname] = norm_data[chname].std(axis=0)

    # one trial per sheet
    ws = wb.active if n == 0 else wb.create_sheet()
    ws.title = op.splitext(op.split(c3dfile)[-1])[0][:31]

    # write some metadata
    _bold_cell(ws, column=1, row=2, value='Trial name:')
    _bold_cell(ws, column=1, row=3, value='N of cycles right:')
    _bold_cell(ws, column=1, row=4, value='N of cycles left:')

    ws.cell(column=2, row=2, value=op.split(c3dfile)[-1])
    ws.cell(column=2, row=3, value=ncycles['L'])
    ws.cell(column=2, row=4, value=ncycles['R'])

    col_headers = [''] + ['frame %d' % k for k in range(101)]

    for col, txt in enumerate(col_headers, 1):
        _bold_cell(ws, column=col, row=5, value=txt)

    # write channel avg and std data
    for _row, chname in enumerate(sorted(chs_valid), 3):
        # averages
        _bold_cell(ws, column=1, row=_row*2, value='%s / average' % chname)
        for col, val in enumerate(avg_data[chname], 2):
            ws.cell(column=col, row=_row*2, value=val)

        _bold_cell(ws, column=1, row=_row*2+1, value='%s / stddev' % chname)
        for col, val in enumerate(std_data[chname], 2):
            ws.cell(column=col, row=_row*2+1, value=val)
    
    _auto_adjust(ws)

wb.save(filename=fname_xls)

